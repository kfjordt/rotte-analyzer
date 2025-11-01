import networkx as nx
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from collections import Counter

import itertools

RAT_COUNT = 10


def read_data(path, sheet_name=0):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
    data = [[cell.value for cell in row] for row in ws.iter_rows()]

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        value = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                data[r - 1][c - 1] = value

    resident_dict = {}
    columns = np.asarray(data).T[1:]
    for column in columns:
        month_label = f"{column[0]} {column[1]}"
        residents = list([resident for resident in column[2:] if resident is not None])
        for resident in residents:
            if resident not in resident_dict:
                resident_dict[resident] = []
            resident_dict[resident].append(month_label)
    return resident_dict


def build_graph(months_by_residents: dict):
    unique_months = sorted(set([m for l in months_by_residents.values() for m in l]))

    residents_by_month = {month: [] for month in unique_months}
    for resident, months in months_by_residents.items():
        for month in months:
            residents_by_month[month].append(resident)

    roomie_graph = {resident: set() for resident in months_by_residents.keys()}
    for month in unique_months:
        residents_in_month = residents_by_month[month]
        for resident in residents_in_month:
            roomie_graph[resident].update(r for r in residents_in_month if r != resident)

    roomie_graph = {r: sorted(list(neighbors)) for r, neighbors in roomie_graph.items()}
    return roomie_graph


def plot_graph(roomie_graph: dict):
    # Build graph
    G = nx.Graph()
    for resident, roommates in roomie_graph.items():
        for roommate in roommates:
            G.add_edge(resident, roommate)

    # Compute stats
    all_paths = dict(nx.all_pairs_shortest_path(G))
    lengths, longest_path, longest_length = [], [], 0
    for u, v in itertools.combinations(G.nodes, 2):
        if v in all_paths[u]:
            path_len = len(all_paths[u][v]) - 1
            lengths.append(path_len)
            if path_len > longest_length:
                longest_length, longest_path = path_len, all_paths[u][v]
    avg_length = sum(lengths) / len(lengths) if lengths else 0

    print(f"Nodes: {G.number_of_nodes()}, Edges: {G.number_of_edges()}")
    print(f"Longest path length: {longest_length}, Path: {longest_path}")
    print(f"Average path length: {avg_length:.2f}")

    pos = nx.kamada_kawai_layout(G)
    degrees = dict(G.degree())
    node_sizes = [400 + 150 * degrees[n] for n in G.nodes()]
    node_colors = [degrees[n] for n in G.nodes()]

    # --- Helper for drawing ---
    def draw(save_path, with_labels=False, fmt="png"):
        plt.figure(figsize=(14, 8), facecolor="white")  # horizontal aspect ratio 8/14
        ax = plt.gca()
        ax.set_facecolor("white")
        nx.draw_networkx_edges(G, pos, alpha=0.25, width=1.0, edge_color="gray")
        nodes = nx.draw_networkx_nodes(
            G,
            pos,
            node_color=node_colors,
            cmap=plt.cm.Oranges,
            node_size=node_sizes,
            edgecolors="black",
            linewidths=0.4,
            alpha=0.95,
        )
        if with_labels:
            nx.draw_networkx_labels(G, pos, font_size=8, font_weight="bold", font_color="red")
        cbar = plt.colorbar(nodes, shrink=0.8, pad=0.02)
        cbar.set_label("Number of Roommates", color="black", fontsize=10)
        plt.setp(plt.getp(cbar.ax.axes, "yticklabels"), color="black")
        plt.title("Roommate Network", fontsize=18, fontweight="bold", color="black", pad=20)
        plt.axis("off")
        plt.tight_layout()
        plt.savefig(save_path, dpi=300, transparent=True, format=fmt)
        plt.close()

    # PowerPoint version (SVG, no labels)
    draw("graph.svg", with_labels=False, fmt="svg")

    # PNG version (with labels)
    draw("graph_labeled.png", with_labels=True, fmt="png")


def plot_box_plot(data: list[float]):
    plt.boxplot(data, vert=True, patch_artist=True)
    plt.title("Box Plot")
    plt.tight_layout()
    plt.ylabel("Values")
    plt.show()
    # plt.savefig("boxplot.pdf")


def plot_histograms(move_in_months, move_out_months):
    # Normalize month abbreviations
    month_map = {
        "Jan": "January",
        "Feb": "February",
        "Mar": "March",
        "Apr": "April",
        "Maj": "May",
        "May": "May",
        "Jun": "June",
        "Jul": "July",
        "Aug": "August",
        "Sep": "September",
        "Okt": "October",
        "Oct": "October",
        "Nov": "November",
        "Dec": "December",
    }

    move_in = [month_map.get(m, m) for m in move_in_months]
    move_out = [month_map.get(m, m) for m in move_out_months]

    months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

    in_counts = Counter(move_in)
    out_counts = Counter(move_out)

    in_values = [in_counts.get(m, 0) for m in months]
    out_values = [out_counts.get(m, 0) for m in months]
    max_count = max(max(in_values), max(out_values))

    fig, axes = plt.subplots(2, 1, figsize=(10, 8), sharex=True, sharey=True)

    axes[0].bar(months, in_values, color="steelblue", alpha=0.7)
    axes[0].set_title("Move In by Month")
    axes[0].set_ylabel("Count")
    axes[0].set_ylim(0, max_count)

    axes[1].bar(months, out_values, color="salmon", alpha=0.7)
    axes[1].set_title("Move Out by Month")
    axes[1].set_ylabel("Count")
    axes[1].set_ylim(0, max_count)

    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()


if __name__ == "__main__":
    # prepare data
    months_by_residents = read_data("rottedata.xlsx")

    # print(months_by_residents)
    roomie_graph = build_graph(months_by_residents)

    # roomie related statistics
    roomie_count_by_resident = {r: len(roomies) for r, roomies in roomie_graph.items()}
    roomie_counts = list(roomie_count_by_resident.values())

    # residency period related statistics
    month_count_by_resident = {r: len(months) for r, months in months_by_residents.items()}
    month_counts = list(month_count_by_resident.values())

    move_in_months = [months[0][-3:] for months in months_by_residents.values()]
    move_out_months = [months[-1][-3:] for months in months_by_residents.values()]


    # for resident, count in sorted(roomie_count_by_resident.items(), key=lambda x: x[1], reverse=True)[:3]:
    #     print(f"{resident}: {count}")
    # print(f"Average roomie count: {np.mean(roomie_counts):.1f}")
    # print(f"Median roomie count: {np.median(roomie_counts):.1f}")
    # print(f"Min residency period: {np.min(month_counts):.1f}")
    # print(f"Max residency period: {np.max(month_counts):.1f}")
    # print(f"Average residency period: {np.mean(month_counts):.1f}")
    # print(f"Median residency period: {np.median(month_counts):.1f}")

    # for resident, count in sorted(month_count_by_resident.items(), key=lambda x: x[1], reverse=True)[:3]:
    #     print(f"{resident}: {count}")

    # plot_histograms(move_in_months, move_out_months)
    # plot_box_plot(month_counts)
    plot_graph(roomie_graph)
